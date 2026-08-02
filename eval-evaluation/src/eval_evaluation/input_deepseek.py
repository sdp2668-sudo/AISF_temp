"""Read the flattened DeepSeek reference workbook."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_HEADERS = ("session_id", "episode_id", "segment_id", "turn_index", "human", "ai")


def load_deepseek_episodes(path: str | Path, sheet_name: str | None = None) -> tuple[list[dict[str, Any]], str]:
    source = Path(path).resolve()
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise ValueError(f"Unable to read DeepSeek XLSX {source}: {exc}") from exc
    try:
        selected = sheet_name or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError(f"DeepSeek worksheet not found: {selected}")
        values = workbook[selected].iter_rows(values_only=True)
        headers = tuple(next(values, ()))
        missing = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing:
            raise ValueError(f"DeepSeek XLSX missing required headers: {', '.join(missing)}")
        rows = []
        for number, values_row in enumerate(values, start=2):
            row = dict(zip(headers, values_row))
            session_id = str(row.get("session_id") or "").strip()
            episode_id = str(row.get("episode_id") or "").strip()
            if not session_id or not episode_id:
                raise ValueError(f"DeepSeek row {number} has empty session_id or episode_id")
            try:
                row["turn_index"] = int(row["turn_index"])
            except (TypeError, ValueError) as exc:
                raise ValueError(f"DeepSeek row {number} has invalid turn_index") from exc
            row["session_id"] = session_id
            row["episode_id"] = episode_id
            row["segment_id"] = str(row.get("segment_id") or "").strip()
            rows.append(row)
    finally:
        workbook.close()

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["session_id"], row["episode_id"])].append(row)
    episodes = []
    for (session_id, episode_id), turns in grouped.items():
        boundaries, ordered_turns = _validate_turns(turns, f"DeepSeek {session_id}/{episode_id}")
        episodes.append({
            "source": "deepseek",
            "session_id": session_id,
            "episode_id": episode_id,
            "start_turn": ordered_turns[0],
            "end_turn": ordered_turns[-1],
            "turns": sorted(turns, key=lambda row: row["turn_index"]),
            "boundaries": boundaries,
            "turn_numbers": ordered_turns,
            "segment_count": len(boundaries) + 1,
        })
    return episodes, selected


def _validate_turns(turns: list[dict[str, Any]], label: str) -> tuple[list[int], list[int]]:
    from .boundaries import boundaries_from_segment_labels

    boundaries, turn_numbers = boundaries_from_segment_labels(turns)
    seen_segments: set[str] = set()
    previous = None
    for row in sorted(turns, key=lambda item: item["turn_index"]):
        segment = row["segment_id"]
        if segment != previous:
            if segment in seen_segments:
                raise ValueError(f"{label} has non-contiguous segment_id {segment}")
            seen_segments.add(segment)
            previous = segment
    return boundaries, turn_numbers
