"""Write auditable JSON and XLSX evaluation artifacts."""

from __future__ import annotations

import json
import os
import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook


def write_artifacts(result: dict[str, Any], qwen_json_path: str | Path, output_dir: str | Path) -> tuple[Path, Path]:
    directory = Path(output_dir).resolve()
    directory.mkdir(parents=True, exist_ok=True)
    stem = Path(qwen_json_path).stem
    json_path = directory / f"{stem}__segment_evaluation.json"
    xlsx_path = directory / f"{stem}__segment_evaluation.xlsx"
    existing = [str(path) for path in (json_path, xlsx_path) if path.exists()]
    if existing:
        raise ValueError(f"Refusing to overwrite existing outputs: {', '.join(existing)}")
    _write_json(result, json_path)
    _write_xlsx(result, xlsx_path)
    return json_path, xlsx_path


def _write_json(value: Any, path: Path) -> None:
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        temp.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temp, path)
    except OSError as exc:
        temp.unlink(missing_ok=True)
        raise ValueError(f"Unable to write JSON {path}: {exc}") from exc


def _write_xlsx(result: dict[str, Any], path: Path) -> None:
    temp = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    workbook = Workbook(write_only=True)
    try:
        _summary_sheet(workbook, result)
        _episode_sheet(workbook, result)
        _boundary_sheet(workbook, result)
        _turn_sheet(workbook, result)
        _observability_sheet(workbook, result)
        workbook.save(temp)
        os.replace(temp, path)
    except (OSError, ValueError) as exc:
        temp.unlink(missing_ok=True)
        raise ValueError(f"Unable to write XLSX {path}: {exc}") from exc


def _summary_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("summary")
    sheet.append(("section", "metric", "value"))
    summary = result["summary"]
    sheet.append(("episodes", "comparable", summary["comparable_episodes"]))
    for mode in ("strict", "loose"):
        for scope in ("micro", "macro"):
            allowed = ("tp", "fp", "fn", "precision", "recall", "f1")
            for metric, value in summary[mode][scope].items():
                if metric not in allowed:
                    continue
                sheet.append((f"{mode}_{scope}", metric, value))
    for name, values in summary["classification"].items():
        sheet.append(("classification", f"{name}_count", values["count"]))
        sheet.append(("classification", f"{name}_rate", values["rate"]))


def _episode_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("episode_metrics")
    headers = (
        "session_id", "qwen_episode_id", "deepseek_episode_id", "start_turn", "end_turn", "turn_count",
        "qwen_segment_count", "deepseek_segment_count", "segment_delta", "classification",
        "strict_tp", "strict_fp", "strict_fn", "strict_precision", "strict_recall", "strict_f1",
        "loose_tp", "loose_fp", "loose_fn", "loose_precision", "loose_recall", "loose_f1",
    )
    sheet.append(headers)
    for episode in result["episodes"]:
        sheet.append(tuple(
            episode.get(name) if name in episode else episode[name.split("_", 1)[0]][name.split("_", 1)[1]]
            for name in headers
        ))


def _boundary_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("boundary_comparison")
    sheet.append((
        "session_id", "qwen_episode_id", "deepseek_episode_id", "mode", "result",
        "qwen_position", "qwen_turn_index", "deepseek_position", "deepseek_turn_index", "offset_turns",
    ))
    for episode in result["episodes"]:
        positions = {index: turn["turn_index"] for index, turn in enumerate(episode["qwen_turns"])}
        for mode in ("strict", "loose"):
            values = episode[mode]
            for match in values["matches"]:
                sheet.append((episode["session_id"], episode["qwen_episode_id"], episode["deepseek_episode_id"], mode, "TP", match["qwen_position"], positions[match["qwen_position"]], match["deepseek_position"], positions[match["deepseek_position"]], match["offset_turns"]))
            for position in values["qwen_unmatched_positions"]:
                sheet.append((episode["session_id"], episode["qwen_episode_id"], episode["deepseek_episode_id"], mode, "FP", position, positions[position], None, None, None))
            for position in values["deepseek_unmatched_positions"]:
                sheet.append((episode["session_id"], episode["qwen_episode_id"], episode["deepseek_episode_id"], mode, "FN", None, None, position, positions[position], None))


def _turn_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("turn_review")
    sheet.append((
        "session_id", "qwen_episode_id", "deepseek_episode_id", "turn_index", "human", "qwen_ai", "deepseek_ai",
        "qwen_segment_id", "deepseek_segment_id", "qwen_boundary_after", "deepseek_boundary_after",
    ))
    for episode in result["episodes"]:
        qwen_by_turn = {row["turn_index"]: row for row in episode["qwen_turns"]}
        baseline_by_turn = {row["turn_index"]: row for row in episode["deepseek_turns"]}
        qwen_boundaries = {item["turn_index"] for item in episode["qwen_boundaries"]}
        baseline_boundaries = {item["turn_index"] for item in episode["deepseek_boundaries"]}
        for turn_index in sorted(qwen_by_turn):
            qwen = qwen_by_turn[turn_index]
            baseline = baseline_by_turn[turn_index]
            sheet.append((
                episode["session_id"], episode["qwen_episode_id"], episode["deepseek_episode_id"], turn_index,
                qwen.get("human") or baseline.get("human") or "", qwen.get("ai") or "", baseline.get("ai") or "",
                qwen["segment_id"], baseline["segment_id"], turn_index in qwen_boundaries, turn_index in baseline_boundaries,
            ))


def _observability_sheet(workbook: Workbook, result: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("run_observability")
    sheet.append(("kind", "session_id", "episode_id", "start_turn", "end_turn", "detail"))
    for key, value in result["run_observability"].items():
        if key in {"excluded_items", "qwen_errors"}:
            continue
        sheet.append(("qwen_run", "", "", "", "", json.dumps({key: value}, ensure_ascii=False)))
    for item in result["run_observability"].get("excluded_items", []):
        sheet.append(("qwen_excluded", item.get("session_id"), item.get("episode_id"), "", "", item.get("reason")))
    for item in result["run_observability"].get("qwen_errors", []):
        sheet.append(("qwen_error", item.get("session_id"), item.get("episode_id"), "", "", item.get("message")))
    for item in result["alignment_excluded"]:
        sheet.append(("alignment_excluded", item.get("session_id"), item.get("episode_id"), item.get("start_turn"), item.get("end_turn"), item.get("reason")))
