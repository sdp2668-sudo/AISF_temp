from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from eval_evaluation.alignment import align_and_evaluate
from eval_evaluation.export import write_artifacts
from eval_evaluation.input_deepseek import load_deepseek_episodes
from eval_evaluation.input_qwen import load_qwen_run, qwen_episodes


class IntegrationTest(unittest.TestCase):
    def test_qwen_json_and_deepseek_workbook_export_all_review_sheets(self):
        qwen_run = {
            "summary": {"episodes": 1, "sessions_failed": 0},
            "errors": [],
            "sessions": [{"session_id": "s1", "status": "succeeded", "episodes": [{
                "episode_id": "e1", "status": "succeeded", "start_turn": 1, "end_turn": 3,
                "segments": [
                    {"segment_id": "s1", "turns": [{"turn_no": 1, "human": "a", "ai": ""}, {"turn_no": 2, "human": "b", "ai": ""}]},
                    {"segment_id": "s2", "turns": [{"turn_no": 3, "human": "c", "ai": ""}]},
                ],
            }]}],
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            qwen_path = root / "qwen.json"
            qwen_path.write_text(json.dumps(qwen_run), encoding="utf-8")
            baseline_path = root / "deepseek.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(("session_id", "episode_id", "segment_id", "turn_index", "human", "ai"))
            sheet.append(("s1", "e1", "s1", 1, "a", ""))
            sheet.append(("s1", "e1", "s1", 2, "b", ""))
            sheet.append(("s1", "e1", "s2", 3, "c", ""))
            workbook.save(baseline_path)

            qwen, excluded = qwen_episodes(load_qwen_run(qwen_path))
            baseline, _ = load_deepseek_episodes(baseline_path)
            episodes, alignment_excluded = align_and_evaluate(qwen, baseline, 1)
            result = {"summary": {"comparable_episodes": 1, "strict": {"micro": episodes[0]["strict"], "macro": {}}, "loose": {"micro": episodes[0]["loose"], "macro": {}}, "classification": {}}, "episodes": episodes, "run_observability": {"excluded_items": excluded, "qwen_errors": []}, "alignment_excluded": alignment_excluded}
            json_path, xlsx_path = write_artifacts(result, qwen_path, root / "results")
            self.assertTrue(json_path.exists())
            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["summary", "episode_metrics", "boundary_comparison", "turn_review", "run_observability"])
            finally:
                workbook.close()
