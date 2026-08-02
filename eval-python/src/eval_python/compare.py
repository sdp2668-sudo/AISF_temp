"""Offline Turn-aligned comparison between Qwen JSON and badcaseOps XLSX."""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .export import SESSION_HEADERS, flatten_qwen_rows, write_json_atomic, write_rows_xlsx_atomic
from .models import ComparisonError


COMPARE_FIELDS = (
    ("episode_id", "episode"),
    ("segment_id", "segment"),
    ("intent_summary", "intent"),
    ("业务", "business"),
    ("子功能", "sub_function"),
    ("AI明确回复不支持问题", "ai_unsupported"),
)

TURN_COMPARISON_HEADERS = (
    "session_id",
    "turn_index",
    "in_baseline",
    "in_qwen",
    "baseline_episode_id",
    "qwen_episode_id",
    "episode_diff",
    "baseline_segment_id",
    "qwen_segment_id",
    "segment_diff",
    "baseline_segment_start",
    "qwen_segment_start",
    "boundary_diff",
    "baseline_intent_summary",
    "qwen_intent_summary",
    "intent_diff",
    "baseline_business",
    "qwen_business",
    "business_diff",
    "baseline_sub_function",
    "qwen_sub_function",
    "sub_function_diff",
    "baseline_ai_unsupported",
    "qwen_ai_unsupported",
    "ai_unsupported_diff",
    "baseline_human",
    "qwen_human",
    "baseline_ai",
    "qwen_ai",
    "any_diff",
)

SESSION_SUMMARY_HEADERS = (
    "session_id",
    "turns_union",
    "turns_only_baseline",
    "turns_only_qwen",
    "episode_diffs",
    "segment_diffs",
    "boundary_diffs",
    "intent_diffs",
    "business_diffs",
    "sub_function_diffs",
    "ai_unsupported_diffs",
    "turns_with_any_diff",
)


def compare_artifacts(
    qwen_json_path: str | Path,
    baseline_xlsx_path: str | Path,
) -> dict[str, Any]:
    qwen_path = Path(qwen_json_path).resolve()
    baseline_path = Path(baseline_xlsx_path).resolve()
    try:
        qwen_result = json.loads(qwen_path.read_text(encoding="utf-8"))
    except OSError as exc:
        raise ComparisonError(f"Unable to read Qwen JSON {qwen_path}: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise ComparisonError(f"Invalid Qwen JSON {qwen_path}: {exc}") from exc
    if not isinstance(qwen_result, dict):
        raise ComparisonError("Qwen result JSON must be an object")

    qwen_rows = flatten_qwen_rows(qwen_result)
    baseline_rows = _load_baseline_rows(baseline_path)
    qwen_map = _row_map(qwen_rows, "Qwen JSON")
    baseline_map = _row_map(baseline_rows, "baseline XLSX")
    qwen_starts = _segment_starts(qwen_map)
    baseline_starts = _segment_starts(baseline_map)

    all_keys = sorted(set(qwen_map) | set(baseline_map), key=lambda key: (key[0], key[1]))
    comparison_rows: list[dict[str, Any]] = []
    for key in all_keys:
        baseline = baseline_map.get(key)
        qwen = qwen_map.get(key)
        row: dict[str, Any] = {
            "session_id": key[0],
            "turn_index": key[1],
            "in_baseline": baseline is not None,
            "in_qwen": qwen is not None,
            "baseline_segment_start": baseline_starts.get(key, False),
            "qwen_segment_start": qwen_starts.get(key, False),
            "baseline_human": _value(baseline, "human"),
            "qwen_human": _value(qwen, "human"),
            "baseline_ai": _value(baseline, "ai"),
            "qwen_ai": _value(qwen, "ai"),
        }
        row["boundary_diff"] = row["baseline_segment_start"] != row["qwen_segment_start"]
        diff_flags = ["boundary_diff"]
        for source_field, output_field in COMPARE_FIELDS:
            baseline_value = _value(baseline, source_field)
            qwen_value = _value(qwen, source_field)
            row[f"baseline_{output_field}"] = baseline_value
            row[f"qwen_{output_field}"] = qwen_value
            flag = f"{output_field}_diff"
            row[flag] = baseline_value != qwen_value
            diff_flags.append(flag)
        row["any_diff"] = (
            baseline is None
            or qwen is None
            or any(bool(row[flag]) for flag in diff_flags)
        )
        comparison_rows.append(row)

    session_rows = _session_summaries(comparison_rows)
    summary = {
        "qwen_json": str(qwen_path),
        "baseline_xlsx": str(baseline_path),
        "sessions": len(session_rows),
        "turns_union": len(comparison_rows),
        "turns_only_baseline": sum(row["in_baseline"] and not row["in_qwen"] for row in comparison_rows),
        "turns_only_qwen": sum(row["in_qwen"] and not row["in_baseline"] for row in comparison_rows),
        "episode_diffs": sum(bool(row["episode_diff"]) for row in comparison_rows),
        "segment_diffs": sum(bool(row["segment_diff"]) for row in comparison_rows),
        "boundary_diffs": sum(bool(row["boundary_diff"]) for row in comparison_rows),
        "intent_diffs": sum(bool(row["intent_diff"]) for row in comparison_rows),
        "business_diffs": sum(bool(row["business_diff"]) for row in comparison_rows),
        "sub_function_diffs": sum(bool(row["sub_function_diff"]) for row in comparison_rows),
        "ai_unsupported_diffs": sum(bool(row["ai_unsupported_diff"]) for row in comparison_rows),
        "turns_with_any_diff": sum(bool(row["any_diff"]) for row in comparison_rows),
    }
    return {"summary": summary, "session_summary": session_rows, "turn_comparison": comparison_rows}


def write_comparison_artifacts(
    comparison: dict[str, Any],
    *,
    output_dir: str | Path,
    qwen_json_path: str | Path,
) -> tuple[Path, Path]:
    directory = Path(output_dir).resolve()
    directory.mkdir(parents=True, exist_ok=True)
    stem = Path(qwen_json_path).stem
    json_path = directory / f"{stem}__comparison.json"
    xlsx_path = directory / f"{stem}__comparison.xlsx"
    existing = [str(path) for path in (json_path, xlsx_path) if path.exists()]
    if existing:
        raise ComparisonError(f"Refusing to overwrite existing comparison: {', '.join(existing)}")
    write_json_atomic(comparison, json_path)
    summary_rows = [{"key": key, "value": value} for key, value in comparison["summary"].items()]
    write_rows_xlsx_atomic(
        [
            ("turn_comparison", TURN_COMPARISON_HEADERS, comparison["turn_comparison"]),
            ("session_summary", SESSION_SUMMARY_HEADERS, comparison["session_summary"]),
            ("summary", ("key", "value"), summary_rows),
        ],
        xlsx_path,
    )
    return json_path, xlsx_path


def _load_baseline_rows(path: Path) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise ComparisonError(f"Unable to read baseline XLSX {path}: {exc}") from exc
    try:
        if "sessions" not in workbook.sheetnames:
            raise ComparisonError("Baseline XLSX must contain sheet 'sessions'")
        values = workbook["sessions"].iter_rows(values_only=True)
        try:
            headers = tuple(next(values))
        except StopIteration as exc:
            raise ComparisonError("Baseline sessions sheet is empty") from exc
        missing = [header for header in SESSION_HEADERS if header not in headers]
        if missing:
            raise ComparisonError(f"Baseline sessions sheet is missing headers: {', '.join(missing)}")
        rows = []
        for values_row in values:
            row = {str(header): value for header, value in zip(headers, values_row)}
            rows.append({header: row.get(header) for header in SESSION_HEADERS})
        return rows
    finally:
        workbook.close()


def _row_map(rows: list[dict[str, Any]], source: str) -> dict[tuple[str, int], dict[str, Any]]:
    result: dict[tuple[str, int], dict[str, Any]] = {}
    for index, row in enumerate(rows, start=2):
        session_id = str(row.get("session_id") or "").strip()
        if not session_id:
            raise ComparisonError(f"{source} row {index} has empty session_id")
        try:
            turn_index = int(row.get("turn_index"))
        except (TypeError, ValueError) as exc:
            raise ComparisonError(f"{source} row {index} has invalid turn_index") from exc
        key = (session_id, turn_index)
        if key in result:
            raise ComparisonError(f"{source} contains duplicate key {session_id}+{turn_index}")
        result[key] = row
    return result


def _segment_starts(rows: dict[tuple[str, int], dict[str, Any]]) -> dict[tuple[str, int], bool]:
    starts: dict[tuple[str, int], bool] = {}
    previous_by_session: dict[str, tuple[int, str] | None] = {}
    for key in sorted(rows, key=lambda value: (value[0], value[1])):
        session_id, turn_index = key
        segment_id = str(rows[key].get("segment_id") or "")
        previous = previous_by_session.get(session_id)
        starts[key] = (
            previous is None
            or turn_index != previous[0] + 1
            or segment_id != previous[1]
        )
        previous_by_session[session_id] = (turn_index, segment_id)
    return starts


def _session_summaries(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row["session_id"]].append(row)
    summaries = []
    for session_id, values in grouped.items():
        summaries.append({
            "session_id": session_id,
            "turns_union": len(values),
            "turns_only_baseline": sum(row["in_baseline"] and not row["in_qwen"] for row in values),
            "turns_only_qwen": sum(row["in_qwen"] and not row["in_baseline"] for row in values),
            "episode_diffs": sum(bool(row["episode_diff"]) for row in values),
            "segment_diffs": sum(bool(row["segment_diff"]) for row in values),
            "boundary_diffs": sum(bool(row["boundary_diff"]) for row in values),
            "intent_diffs": sum(bool(row["intent_diff"]) for row in values),
            "business_diffs": sum(bool(row["business_diff"]) for row in values),
            "sub_function_diffs": sum(bool(row["sub_function_diff"]) for row in values),
            "ai_unsupported_diffs": sum(bool(row["ai_unsupported_diff"]) for row in values),
            "turns_with_any_diff": sum(bool(row["any_diff"]) for row in values),
        })
    return summaries


def _value(row: dict[str, Any] | None, field: str) -> Any:
    if row is None:
        return ""
    value = row.get(field)
    return "" if value is None else value
