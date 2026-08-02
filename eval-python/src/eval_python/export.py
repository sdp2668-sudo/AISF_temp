"""Atomic JSON and XLSX result serialization."""

from __future__ import annotations

import json
import os
import re
import uuid
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from .models import ExportError


SESSION_HEADERS = (
    "session_id",
    "episode_id",
    "segment_id",
    "turn_index",
    "total_turns",
    "time",
    "intent_summary",
    "业务",
    "子功能",
    "human",
    "rewritten_query",
    "ai",
    "AI明确回复不支持问题",
)

PERFORMANCE_HEADERS = (
    "stage",
    "session_id",
    "episode_id",
    "segment_id",
    "elapsed_seconds",
    "attempts",
    "prompt_tokens",
    "completion_tokens",
    "total_tokens",
)

ERROR_HEADERS = (
    "stage",
    "session_id",
    "episode_id",
    "segment_id",
    "error_type",
    "message",
)


def write_run_artifacts(
    result: dict[str, Any],
    *,
    input_path: str | Path,
    output_dir: str | Path,
) -> tuple[Path, Path]:
    directory = Path(output_dir).resolve()
    directory.mkdir(parents=True, exist_ok=True)
    model = _filename_part(str(result.get("model", {}).get("name") or "qwen"))
    stem = _filename_part(Path(input_path).stem)
    run_id = _filename_part(str(result.get("run_id") or uuid.uuid4().hex[:8]))
    base = f"{stem}__{model}__{run_id}"
    json_path = directory / f"{base}.json"
    xlsx_path = directory / f"{base}.xlsx"
    _ensure_absent(json_path, xlsx_path)
    write_json_atomic(result, json_path)
    write_result_xlsx_atomic(result, xlsx_path)
    return json_path, xlsx_path


def flatten_qwen_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    sessions = result.get("sessions", [])
    if not isinstance(sessions, list):
        raise ExportError("Result sessions must be an array")
    for session in sessions:
        if not isinstance(session, dict):
            raise ExportError("Result Session must be an object")
        total_turns = int(session.get("turn_count") or 0)
        for episode in session.get("episodes", []):
            for segment in episode.get("segments", []):
                for turn in segment.get("turns", []):
                    rows.append({
                        "session_id": session.get("session_id", ""),
                        "episode_id": episode.get("episode_id", ""),
                        "segment_id": segment.get("segment_id", ""),
                        "turn_index": turn.get("turn_no", 0),
                        "total_turns": total_turns,
                        "time": turn.get("timestamp") or "",
                        "intent_summary": segment.get("intent_summary") or "",
                        "业务": segment.get("business") or "",
                        "子功能": segment.get("sub_function") or "",
                        "human": turn.get("human") or "",
                        "rewritten_query": turn.get("rewrite_query") or "",
                        "ai": turn.get("ai") or "",
                        "AI明确回复不支持问题": segment.get("ai_unsupported") or "",
                    })
    return rows


def write_json_atomic(value: Any, path: str | Path) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp = target.with_name(f".{target.name}.{uuid.uuid4().hex}.tmp")
    try:
        temp.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temp, target)
    except OSError as exc:
        try:
            temp.unlink(missing_ok=True)
        except OSError:
            pass
        raise ExportError(f"Unable to write JSON {target}: {exc}") from exc


def write_result_xlsx_atomic(result: dict[str, Any], path: str | Path) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp = target.with_name(f".{target.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    workbook = Workbook(write_only=True)
    try:
        sessions_sheet = workbook.create_sheet("sessions")
        sessions_sheet.append(SESSION_HEADERS)
        for row in flatten_qwen_rows(result):
            sessions_sheet.append(tuple(_cell(sessions_sheet, row.get(header)) for header in SESSION_HEADERS))

        summary_sheet = workbook.create_sheet("run_summary")
        summary_sheet.append(("key", "value"))
        summary_values = {
            "schema_version": result.get("schema_version"),
            "source_baseline": result.get("source_baseline"),
            "run_id": result.get("run_id"),
            "input_file": result.get("input_file"),
            "started_at": result.get("started_at"),
            "completed_at": result.get("completed_at"),
            "elapsed_seconds": result.get("elapsed_seconds"),
            "model": result.get("model"),
            "features": result.get("features"),
            "filters": result.get("filters"),
            "summary": result.get("summary"),
        }
        for key, value in summary_values.items():
            summary_sheet.append((_cell(summary_sheet, key), _cell(summary_sheet, _display_value(value))))

        performance_sheet = workbook.create_sheet("performance")
        performance_sheet.append(PERFORMANCE_HEADERS)
        for metric in result.get("metrics", []):
            performance_sheet.append(tuple(
                _cell(performance_sheet, metric.get(header)) for header in PERFORMANCE_HEADERS
            ))

        errors_sheet = workbook.create_sheet("errors")
        errors_sheet.append(ERROR_HEADERS)
        for error in result.get("errors", []):
            errors_sheet.append(tuple(_cell(errors_sheet, error.get(header)) for header in ERROR_HEADERS))

        workbook.save(temp)
        os.replace(temp, target)
    except (OSError, ValueError) as exc:
        try:
            temp.unlink(missing_ok=True)
        except OSError:
            pass
        raise ExportError(f"Unable to write XLSX {target}: {exc}") from exc


def write_rows_xlsx_atomic(
    sheets: Iterable[tuple[str, tuple[str, ...], list[dict[str, Any]]]],
    path: str | Path,
) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp = target.with_name(f".{target.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    workbook = Workbook(write_only=True)
    try:
        for title, headers, rows in sheets:
            sheet = workbook.create_sheet(title)
            sheet.append(headers)
            for row in rows:
                sheet.append(tuple(_cell(sheet, row.get(header)) for header in headers))
        workbook.save(temp)
        os.replace(temp, target)
    except (OSError, ValueError) as exc:
        try:
            temp.unlink(missing_ok=True)
        except OSError:
            pass
        raise ExportError(f"Unable to write XLSX {target}: {exc}") from exc


def _cell(sheet, value: Any):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    cell = WriteOnlyCell(sheet, value=text)
    cell.data_type = "s"
    return cell


def _display_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-.")
    return cleaned or "result"


def _ensure_absent(*paths: Path) -> None:
    existing = [str(path) for path in paths if path.exists()]
    if existing:
        raise ExportError(f"Refusing to overwrite existing output: {', '.join(existing)}")

