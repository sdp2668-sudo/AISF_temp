from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from eval_python.compare import TURN_COMPARISON_HEADERS, compare_artifacts, write_comparison_artifacts
from eval_python.export import SESSION_HEADERS, flatten_qwen_rows, write_json_atomic, write_result_xlsx_atomic, write_rows_xlsx_atomic


def qwen_result() -> dict:
    return {
        "schema_version": "1.0",
        "source_baseline": "test",
        "run_id": "run-1",
        "input_file": "sessions.json",
        "started_at": "2026-07-27T00:00:00Z",
        "completed_at": "2026-07-27T00:00:01Z",
        "elapsed_seconds": 1.0,
        "model": {"name": "Qwen3-32B"},
        "features": {"enable_subscene": True, "enable_ai_unsupported": True},
        "filters": {"excluded_items": []},
        "summary": {"turn_rows": 2, "errors": 0},
        "metrics": [],
        "errors": [],
        "sessions": [{
            "session_id": "s1",
            "turn_count": 2,
            "episodes": [{
                "episode_id": "e1",
                "segments": [{
                    "segment_id": "s1",
                    "intent_summary": "播放音乐",
                    "business": "音乐",
                    "sub_function": "搜索",
                    "ai_unsupported": "否",
                    "turns": [{
                        "turn_no": 1,
                        "human": "播放青花瓷",
                        "ai": "好的",
                        "rewrite_query": "播放歌曲青花瓷",
                        "timestamp": "2026-07-27T10:00:00+08:00",
                    }, {
                        "turn_no": 2,
                        "human": "暂停",
                        "ai": "已暂停",
                        "rewrite_query": None,
                        "timestamp": "2026-07-27T10:01:00+08:00",
                    }],
                }],
            }],
        }],
    }


class ExportCompareTest(unittest.TestCase):
    def test_flatten_and_workbook_keep_exact_sessions_contract(self):
        result = qwen_result()
        rows = flatten_qwen_rows(result)
        self.assertEqual(tuple(rows[0]), SESSION_HEADERS)
        self.assertEqual(rows[0]["rewritten_query"], "播放歌曲青花瓷")

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            write_result_xlsx_atomic(result, path)
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["sessions", "run_summary", "performance", "errors"])
                session_rows = list(workbook["sessions"].iter_rows(values_only=True))
                self.assertEqual(tuple(session_rows[0]), SESSION_HEADERS)
                self.assertEqual(len(session_rows), 3)
            finally:
                workbook.close()

    def test_offline_compare_exports_value_and_boundary_flags(self):
        result = qwen_result()
        baseline_rows = [
            dict(zip(SESSION_HEADERS, (
                "s1", "e1", "s1", 1, 2, "2026-07-27T10:00:00+08:00",
                "播放歌曲", "音乐", "搜索", "播放青花瓷", "播放歌曲青花瓷", "好的", "否",
            ))),
            dict(zip(SESSION_HEADERS, (
                "s1", "e2", "s2", 2, 2, "2026-07-27T10:01:00+08:00",
                "控制播放", "系统操控", "播放控制", "暂停", "", "已暂停", "否",
            ))),
        ]

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            qwen_path = root / "qwen.json"
            baseline_path = root / "baseline.xlsx"
            write_json_atomic(result, qwen_path)
            write_rows_xlsx_atomic([("sessions", SESSION_HEADERS, baseline_rows)], baseline_path)

            comparison = compare_artifacts(qwen_path, baseline_path)
            second = comparison["turn_comparison"][1]
            self.assertTrue(second["episode_diff"])
            self.assertTrue(second["segment_diff"])
            self.assertTrue(second["boundary_diff"])
            self.assertEqual(comparison["summary"]["episode_diffs"], 1)
            self.assertEqual(comparison["summary"]["segment_diffs"], 1)

            json_path, xlsx_path = write_comparison_artifacts(
                comparison,
                output_dir=root / "comparison",
                qwen_json_path=qwen_path,
            )
            exported = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual(exported["summary"]["turns_union"], 2)
            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            try:
                headers = tuple(next(workbook["turn_comparison"].iter_rows(values_only=True)))
                self.assertEqual(headers, TURN_COMPARISON_HEADERS)
                self.assertIn("episode_diff", headers)
                self.assertIn("segment_diff", headers)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
